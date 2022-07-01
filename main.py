from PySide6.QtCore import *
from PySide6.QtGui import *
from PySide6.QtWidgets import *
import modules.pyApp as pyApp
import sys
import openpyxl
from pathlib import Path
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
from typing import *


# 获取Excel路径，并修改软件的工作目录
def getExcelFilePath(currentWorkDir='./') -> str:
    excelFile = QFileDialog.getOpenFileName(
        MainWindow, "选择excel文件", currentWorkDir, 'Excel Files (*.xlsx *.xls)')  # 返回选中的路径形成的列表
    excelFilePath = excelFile[0]  # 取出路径列表中存储的路径

    global EXCEL_FILE_DIR
    EXCEL_FILE_DIR = str(Path(excelFilePath).parent)  # 重新设置工作目录
    global EXCEL_FLAG
    if Path(excelFilePath).suffix == '.xls':  # 重新设置excel标识
        EXCEL_FLAG = 0
    else:
        EXCEL_FLAG = 1

    return excelFilePath


# 获取Excel的sheet清单
def getExcelSheetName(path: str) -> list:
    wb = openpyxl.load_workbook(Path(path))
    wbSheetsList = [i.title for i in wb]
    wb.close()
    return wbSheetsList


# 定义tableWidget控件的排版方法，此控件9*6
def gridTableWidget(list: list) -> None:

    def _ListIter(list):
        for i in list:
            yield i

    listIter = _ListIter(list)
    ui.tableWidget.clearContents()
    itemColumn = 0

    while itemColumn < ui.tableWidget.columnCount():
        itemRow = 0
        while itemRow < ui.tableWidget.rowCount():
            try:
                ui.tableWidget.setItem(
                    itemRow, itemColumn, QTableWidgetItem(next(listIter)))
                itemRow += 1
            except StopIteration:
                return None
        itemColumn += 1


# 定义获取表格选定内容的方法
def getTabArray():
    return [i.text() for i in ui.tableWidget.selectedItems()]


# 定义刷新表格区域的方法
def tabRefush():
    try:
        excelFilePath = ui.lineEdit.text()  # 读取lineEdit存储的excel路径
        wbSheetsList = getExcelSheetName(excelFilePath)
        ui.tableWidget.setColumnCount(len(wbSheetsList)//9+1)
        gridTableWidget(wbSheetsList)  # 将sheet清单写入tableWidget
    except:
        pass
    finally:
        pass
    return None


# 定义openFile按钮的动作
def cmdOpenExcelFile() -> None:
    excelFilePath = getExcelFilePath(EXCEL_FILE_DIR)  # 获取excel路径
    if excelFilePath != '':
        ui.lineEdit.setText(excelFilePath)  # 将excel路径写入lineEdit
    else:
        return None

    try:
        wbSheetsList = getExcelSheetName(
            excelFilePath)  # 获取excel的sheet清单
        ui.tableWidget.setColumnCount(
            len(wbSheetsList)//9+1)  # 根据sheet清单尺寸决定表格尺寸
        gridTableWidget(wbSheetsList)  # 将sheet清单写入tableWidget
    except:
        return None


# 定义commitFileCMD按钮的动作
def cmdCommitFile() -> None:
    try:
        font = Font(underline='single', color='FF0000FF')
        try:
            excelFilePath = ui.lineEdit.text()  # 读取lineEdit存储的excel路径
            wb = openpyxl.load_workbook(Path(excelFilePath))  # 打开文件
        except:
            QMessageBox.critical(MainWindow, '错误', '未选择表格！')
            return None

        if getTabArray():
            wsWorkSheetList = getTabArray()
        else:
            wsWorkSheetList = [i.title for i in wb if i.title != '目录']

        if '目录' not in [i.title for i in wb]:  # 建立空的目录sheet
            wb.create_sheet('目录', 0)
        else:
            wb['目录'].delete_cols(1, 2)
        wsList = wb['目录']

        rownum, colnum = 1, 1
        wsList.cell(row=rownum, column=colnum, value='目 录')
        for i in wsWorkSheetList:
            rownum += 1
            wsList.cell(row=rownum, column=colnum, value=i).font = font
            wsList.cell(row=rownum, column=colnum).hyperlink = Hyperlink(
                ref='', location='\'%s\'!A1' % i, tooltip=None, display='%s' % i, id=None)

            wb[i]['A3'].font = font
            wb[i]['A3'].hyperlink = Hyperlink(
                ref='', location='\'目录\'!A1', tooltip=None, display='目录', id=None)
        wb.save(excelFilePath)
        wb.close()
        getTabArray()
        tabRefush()
        QMessageBox.information(MainWindow, '信息', '建立成功！')
    except:
        QMessageBox.critical(MainWindow, '错误', '未知错误！')
    finally:
        pass


# 主程序
if __name__ == "__main__":
    try:
        APP_VERSION = 'v0.3.0'  # 软件版本号
        EXCEL_FILE_DIR = './'  # 默认的工作目录
        EXCEL_FLAG = None  # excel文件类型的标识，1标识xlsx，0标识xls，None表示未取得excel文件

        app = QApplication(sys.argv)
        MainWindow = QMainWindow()

        ui = pyApp.Ui_MainWindow()
        ui.setupUi(MainWindow)
        ui.stackedWidget.setCurrentIndex(0)
        ui.label_8.setText(APP_VERSION + '  ')

        MainWindow.show()

        ui.openFilesButton.clicked.connect(
            cmdOpenExcelFile)  # 监控openFilesButton的click动作
        ui.commitButton.clicked.connect(
            cmdCommitFile)  # 监控commitButton的click动作
        ui.toggleButton.clicked.connect(
            lambda: ui.stackedWidget.setCurrentIndex(0))  # 监控toggleButton的click动作
        ui.btn_home.clicked.connect(
            lambda: ui.stackedWidget.setCurrentIndex(1))  # 监控btn_home的click动作
        ui.btn_widgets.clicked.connect(
            lambda: ui.stackedWidget.setCurrentIndex(2))  # 监控功能2的click动作
        ui.btn_new.clicked.connect(
            lambda: ui.stackedWidget.setCurrentIndex(3))  # 监控功能3的click动作
        ui.btn_save.clicked.connect(
            lambda: ui.stackedWidget.setCurrentIndex(4))  # 监控功能4的click动作
        ui.btn_exit.clicked.connect(
            lambda: ui.stackedWidget.setCurrentIndex(5))  # 监控功能5的click动作

        sys.exit(app.exec())

    finally:
        pass
