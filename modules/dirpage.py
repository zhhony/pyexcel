from cv2 import findNonZero
import modules.appui as appui
import openpyxl
import xlrd
import win32com.client as client
import threading
from PySide6.QtCore import *
from PySide6.QtGui import *
from PySide6.QtWidgets import *
from pathlib import Path
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font


class DirPage():
    def __init__(self, MainWindow: QMainWindow, ui: appui.Ui_MainWindow) -> None:
        self._EXCEL_FILE_DIR = './'  # 默认的工作目录
        self._EXCEL_FLAG = None  # excel文件类型的标识，1标识xlsx，0标识xls，None表示未取得excel文件
        self._FONT = Font(underline='single', color='FF0000FF')
        self._MainWindow = MainWindow
        self._ui = ui
        self.lock = threading.Lock()
        self.excelSheetName = []
        self.excelFilePath = None

    # 判断excel是否为xlsx文件
    def isXlsx(self, excelFilePath: str) -> bool:
        if Path(excelFilePath).suffix == '.xlsx':
            self._EXCEL_FLAG = 1
            return True
        elif Path(excelFilePath).suffix == '.xls':
            self._EXCEL_FLAG = 0
            return False
        else:
            self._EXCEL_FLAG = None
            return None

    # xls转xlsxtemp
    def transXlsToXlsx(self, excelFilePath: str) -> None:
        try:
            excel = client.gencache.EnsureDispatch('Excel.Application')
            wbTemp = excel.Workbooks.Open(excelFilePath)
            wbTemp.SaveAs(excelFilePath + "xtemp", FileFormat=51)
        except Exception as e:
            print(e.args)
        finally:
            wbTemp.Close(True)
            excel.Application.Quit()
            excel.Quit

    # xlsxtemp转xls
    def transXlsxToXls(self, excelFilePath: str) -> None:
        try:
            excel = client.gencache.EnsureDispatch('Excel.Application')
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(excelFilePath)
            wb.SaveAs(excelFilePath[:-5], FileFormat=56)
        finally:
            wb.Close()
            excel.Application.Quit()

    # 获取Excel路径,并赋给变量工作目录
    def getExcelFilePath(self) -> str:
        excelFile = QFileDialog.getOpenFileName(
            self._MainWindow, "选择excel文件", self._EXCEL_FILE_DIR, 'Excel Files (*.xlsx *.xls)')  # 返回选中的路径形成的列表
        self.isXlsx(excelFile[0])
        self._EXCEL_FILE_DIR = str(Path(excelFile[0]).parent)
        return excelFile[0]

    # 获取Excel的sheet清单
    def getExcelSheetName(self, path: str) -> list:
        if self._EXCEL_FLAG == None:
            print('获取excel的sheet清单无效,未选择excel对象')
            return []
        with open(path, 'rb') as f:
            if self._EXCEL_FLAG:
                wb = openpyxl.load_workbook(f)
                self.excelSheetName = [i.title for i in wb]
                wb.close()

            else:
                wb = xlrd.open_workbook(file_contents=f.read(), on_demand=True)
                self.excelSheetName = wb.sheet_names()
                wb.release_resources()

    # 定义tableWidget控件的排版方法，此控件9*6
    def gridTableWidget(self, list: list) -> None:

        def _ListIter(list):
            for i in list:
                yield i
        listIter = _ListIter(list)
        self._ui.tableWidget.clearContents()
        itemColumn = 0

        while itemColumn < self._ui.tableWidget.columnCount():
            itemRow = 0
            while itemRow < self._ui.tableWidget.rowCount():
                try:
                    self._ui.tableWidget.setItem(
                        itemRow, itemColumn, QTableWidgetItem(next(listIter)))
                    itemRow += 1
                except StopIteration:
                    return None
            itemColumn += 1

    # 定义获取tableWidget控件选定内容的方法
    def getTableWidgetArray(self) -> list:
        return [i.text() for i in self._ui.tableWidget.selectedItems()]

    # 定义刷新表格区域的方法
    def refTableWidget(self) -> None:
        try:
            # 读取lineEdit存储的excel路径
            self.getExcelSheetName(self.excelFilePath)
            self._ui.tableWidget.setColumnCount(len(self.excelSheetName)//9+1)
            # 将sheet清单写入tableWidget
            self.gridTableWidget(self.excelSheetName)
        except Exception as e:
            print(e)
        finally:
            pass

    # 建立超链接
    def setHyperlink(self, excelFilePath, excelSheetName, wsWorkSheetList):
        if not self._EXCEL_FLAG:
            self.transXlsToXlsx(excelFilePath)  # 将文件转化为xlsx
            excelFilePath = excelFilePath + 'xtemp'

        with open(excelFilePath, 'rb') as f:
            wb = openpyxl.load_workbook(f)

            if '目录' not in excelSheetName:  # 建立空的目录sheet
                wb.create_sheet('目录', 0)
            wb['目录'].delete_cols(1, 2)
            wsList = wb['目录']

            rownum, colnum = 1, 1
            wsList.cell(row=rownum, column=colnum, value='目 录')

            for i in wsWorkSheetList:
                rownum += 1
                wsList.cell(row=rownum, column=colnum,
                            value=i).font = self._FONT
                wsList.cell(row=rownum, column=colnum).hyperlink = Hyperlink(
                    ref='', location='\'%s\'!A1' % i, tooltip=None, display='%s' % i, id=None)

                wb[i]['A3'].font = self._FONT
                wb[i]['A3'].hyperlink = Hyperlink(
                    ref='', location='\'目录\'!A1', tooltip=None, display='目录', id=None)
            wb.save(excelFilePath)
            wb.close()

        if not self._EXCEL_FLAG:
            self.transXlsxToXls(excelFilePath)  # 将文件转化为xls
            Path(excelFilePath).unlink()

    # 定义openFile按钮的动作
    def cmdOpenExcelFile(self) -> None:
        self.excelFilePath = self.getExcelFilePath()  # 获取excel路径
        try:
            if self.excelFilePath != '':
                self._ui.lineEdit.setText(
                    self.excelFilePath)  # 将excel路径写入lineEdit

                # 起一个子线程获取清单
                Th1 = threading.Thread(target=self.getExcelSheetName, args=(
                    self.excelFilePath,), name='读取sheet清单')
                Th1.start()

                # 主线程阻塞
                self._MainWindow.setCursor(QCursor(Qt.BusyCursor))
                print('a')
                Th1.join()
                print('b')
            else:
                return None

        finally:
            self.refTableWidget()
            self._MainWindow.setCursor(QCursor(Qt.ArrowCursor))

    # 定义commitFileCMD按钮的动作
    def cmdCommitFile(self) -> None:
        if self._EXCEL_FLAG == None:
            print("未选择任何excel对象")
            return None

        self.lock.acquire()

        wsWorkSheetList = self.getTableWidgetArray() if self.getTableWidgetArray() != [
        ] else [i for i in self.excelSheetName if i != '目录']  # 存储需要建立目录的sheet列表,优先把用户选择的sheet赋值给wsWorkSheetList

        try:
            # 起一个子线程建立超链接
            Th1 = threading.Thread(target=self.setHyperlink, args=(
                self.excelFilePath, self.excelSheetName, wsWorkSheetList), name='超链接建立线程')
            Th1.start()

            # 主线程阻塞
            self._MainWindow.setCursor(QCursor(Qt.BusyCursor))
            Th1.join()
            self.refTableWidget()
            self._MainWindow.setCursor(QCursor(Qt.ArrowCursor))

        finally:
            self.lock.release()
            QMessageBox.information(self._MainWindow, '信息', '建立成功！')
