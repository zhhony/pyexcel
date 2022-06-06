import openpyxl
from pathlib import Path
import argparse
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
from typing import *
from win32com import client


# 验证path是否是一个合法的文件路径
def PathVerify(userPath: str) -> bool:
    if Path(userPath).exists():
        if Path(userPath).is_file():
            return True
    else:
        return False


# 验证path是否是xls、xlsx文件
def TypeVerify(userPath: str) -> bool:
    if Path(userPath).match('*.xlsx') or Path(userPath).match('*.xls'):
        return True
    else:
        return False


# xls转xlsx
def TypeTransform(userPath: str) -> str:
    excel = client.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(userPath)
    wb.SaveAs(userPath + "x", FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    return userPath + "x"


# 定义异常类
class ValueError(Exception):
    pass


# 主程序
if __name__ == "__main__":
    try:
        parser = argparse.ArgumentParser()

        parser.add_argument('-p', '--path', help='excel所在的路径，默认放在当前用户文件夹下面')
        args = parser.parse_args()

        path = args.path

        # 路径合法性校验
        if PathVerify(path) == False:
            print('不是一个合法的路径，请检查字符串两边是否有引号')
            raise ValueError
        if TypeVerify(path) == False:
            print('不是一个合法的Excel文件，必须是xls或xlsx类型的excel文件')
            raise ValueError

        # xls转xlsx
        if Path(path).match('*.xls'):
            userCmd = None
            while userCmd == None:
                userCmd = input('是否将xls文件转为xlsx文件？请输入yes[Y] or no[N]:\n')
                try:
                    if userCmd not in {'Y', 'N'}:
                        userCmd = None
                        raise ValueError
                except ValueError:
                    print('不是一个有效的输入结果')
                    continue
            if userCmd == 'Y':
                path = TypeTransform(path)
            else:
                raise ValueError

        font = Font(underline='single', color='FF0000FF')
        wb = openpyxl.load_workbook(Path(path))

        if '目录' not in [i.title for i in wb]:
            wb.create_sheet('目录', 0)
        else:
            wb['目录'].delete_cols(1, 2)
        wsList = wb['目录']

        rownum, colnum = 1, 1
        wsWorkSheetList = [i.title for i in wb if i.title != '目录']
        wsList.cell(row=rownum, column=colnum, value='目 录')
        for i in wsWorkSheetList:
            rownum += 1
            wsList.cell(row=rownum, column=colnum, value=i).font = font
            wsList.cell(row=rownum, column=colnum).hyperlink = Hyperlink(
                ref='', location='\'%s\'!A1' % i, tooltip=None, display='%s' % i, id=None)

            wb[i]['A3'].font = font
            wb[i]['A3'].hyperlink = Hyperlink(
                ref='', location='\'目录\'!A1', tooltip=None, display='目录', id=None)
        wb.save(path)
        wb.close()
    except ValueError:
        print('程序终止')
# C:\Users\zhhony\Desktop\titanic.xlsx
